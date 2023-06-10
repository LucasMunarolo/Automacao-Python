import openpyxl
import os
from datetime import date
import win32com.client as win32


# Define o vencimento conforme a data de hoje
def define_vencimento(vencimento):
    # Data de hoje
    data_hoje = date.today()
    # Dia de hoje
    dia_hoje = data_hoje.day
    # Se o dia de vencimento for maior que o dia de hoje, o mês de pagamento é o mês atual
    if vencimento > dia_hoje:
        mes = data_hoje.month
        # E o ano é o mesmo
        ano = data_hoje.year
    # Se o dia de vencimento for menor que o dia de hoje, o mês de pagamento é o próximo mês
    else:
        mes_atual = data_hoje.month
        # Se for qualquer mês até novembro, adiciona 1 para retornar o próximo mês
        if mes_atual < 12:
            mes = data_hoje.month + 1
            # E o ano é o mesmo
            ano = data_hoje.year
        else:
            # Se for dezembro, o próximo mês é janeiro
            mes = 1
            # E o ano será o próximo
            ano = data_hoje.year + 1
    # Formata a data de vencimento
    data_vencimento = f'{vencimento}/{mes}/{ano}'
    # Retorna a data de vencimento
    return data_vencimento


# Validar anexos que serão enviados no e-mail
def validar_anexo(anexo):
    # De acordo com as regras de negócio, só serão permitidos anexos em formato .xlsx ou .pdf
    if not anexo.endswith('.pdf') or not anexo.endswith('.xlsx'):
        raise ValueError('Os anexos precisam estar no formato .pdf ou .xlsx')
    else:
        return anexo


class AutomacaoEnvioPlanilhas:

    def __init__(self, nome_planilha_contatos, caminho_pastas):
        self._nome_planilha_contatos = nome_planilha_contatos
        self._caminho_pastas = caminho_pastas
        self._criar_pastas_empresas()

    @property
    def _planilha_contatos(self):
        # Abre a pasta de trabalho do Excel
        pasta_trabalho = openpyxl.load_workbook(self._nome_planilha_contatos)
        # Guarda o nome da planilha em uma variável
        nome_planilha = pasta_trabalho.get_sheet_names()[0]
        # Seleciona a planilha que será utilizada
        planilha_contatos = pasta_trabalho[nome_planilha]
        # Retorna o objeto do tipo WorkSheet
        return planilha_contatos

    def _criar_pastas_empresas(self):
        # Verifica os arquivos da pasta onde serão inseridas as pastas da empresa
        arquivos_diretorio = os.listdir(self._caminho_pastas)
        # Percorre as linhas da planilha de contatos
        for linha in range(2, self._planilha_contatos.max_row + 1):
            # Grava o nome da empresa em uma variável
            nome_empresa = self._planilha_contatos['A' + str(linha)].value
            # Se a empresa não tiver uma pasta, cria uma nova pasta
            if nome_empresa not in arquivos_diretorio:
                # Caminho onde será criada a pasta com o nome da empresa
                caminho_pasta_nome_empresa = f'{self._caminho_pastas}\\{nome_empresa}'
                # Cria a pasta da empresa
                os.makedirs(caminho_pasta_nome_empresa)

    # Retornar data de pagamento, através da planilha de contatos, no formato dd/MM/aaaa
    def _retornar_data_pagamento(self, nome_empresa):
        # Percorre as linhas da planilha de contatos:
        for linha in range(2, self._planilha_contatos.max_row + 1):
            # Verifica se o nome da empresa está na linha
            if self._planilha_contatos.cell(row=linha, column=1).value.strip() == nome_empresa.strip():
                # Se encontrar, verifica a data de vencimento desta empresa
                vencimento = self._planilha_contatos.cell(row=linha, column=2).value
                # Retorna a data de vencimento formatada
                return define_vencimento(vencimento)
        # Se não encontrar retorna uma excessão
        raise Exception('Empresa não encontrada!')

    # Função para encontrar o destinatário do e-mail
    def _retornar_email_destinatario(self, nome_empresa):
        # Percorre as linhas da planilha de contatos:
        for linha in range(2, self._planilha_contatos.max_row + 1):
            # Procura o parâmetro nome_empresa na primeira coluna da planilha de contatos
            if self._planilha_contatos.cell(row=linha, column=1).value.strip() == nome_empresa.strip():
                # Se encontrar, verifica os e-mails da empresa
                email = self._planilha_contatos.cell(row=linha, column=3).value
                # Retorna os e-mails da empresa
                return email
        # Se não encontrar retorna uma excessão
        raise Exception('Empresa não encontrada!')

    # Retorna uma lista contendo o nome de todas as empresas
    def retornar_todas_empresas(self):
        # Instancia uma lista vazia onde serão adicionadas as empresas
        lista_empresas = []
        # Percorre as linhas da planilha de contatos
        for linha in range(2, self._planilha_contatos.max_row + 1):
            # Obtém o nome da empresa
            nome_empresa = self._planilha_contatos.cell(row=linha, column=1).value
            # Adiciona esta empresa na lista de empresas
            lista_empresas.append(nome_empresa)
        # Retorna a lista contendo todas as empresas
        return lista_empresas

    # Função para enviar um e-mail personalizado para cada empresa
    def enviar_email(self, nome_empresa, anexo_1, anexo_2=None, anexo_3=None):
        # Abre a aplicação do outlook
        outlook = win32.Dispatch('outlook.application')
        # Cria um e-mail
        email = outlook.CreateItem(0)
        # Informações necessárias para o envio do e-mail
        nome_empresa_formatado = nome_empresa.title()
        data_pagamento = self._retornar_data_pagamento(nome_empresa)
        email_empresa = self._retornar_email_destinatario(nome_empresa)
        assunto_email = f'{nome_empresa_formatado} - Parcelas com vencimento {data_pagamento}'
        texto_email = f'''
        <p>Olá {nome_empresa_formatado}!</p>
        
        <p>Segue em anexo informações referentes às parcelas com vencimento em
        {data_pagamento}.</p>
        
        <p>Qualquer dúvida ficamos à disposição</p>
        
        <p>Atenciosamente</p>
        '''
        # Configurando as informações do e-mail
        email.To = email_empresa
        email.Subject = assunto_email
        email.HTMLBody = texto_email
        # Adicionando os anexos válidos ao e-mail
        anexos = [anexo_1, anexo_2, anexo_3]
        for anexo in anexos:
            if anexo is not None:
                anexo_valido = validar_anexo(anexo)
                email.Attachments.Add(anexo_valido)
        # Enviando o e-mail
        email.Send()

