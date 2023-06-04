import os
import pandas as pd
import openpyxl
from datetime import date
from openpyxl.utils.cell import get_column_letter
from openpyxl.utils.cell import column_index_from_string


class OrganizadorPlanilha:

    def __init__(self, caminho_planilha, nome_empresa):
        self._caminho_planilha = caminho_planilha
        self._validar_planilha()  # Verifica se o caminho da planilha é válido
        self._nome_empresa = nome_empresa
        self._converter_csv_para_xlsx()  # Converte a planilha para .xlsx se estiver em formato .csv
        self._pasta_trabalho, self._planilha = self._abrir_pasta_trabalho()

    # Função valida se o caminho da planilha é válido (usada no construtor)
    def _validar_planilha(self):
        # Se o caminho não for de uma planilha em formato .csv ou .xlsx, retorna um ValueError
        if not self._caminho_planilha.endswith('.csv') and not self._caminho_planilha.endswith('.xlsx'):
            raise ValueError('A aplicação só aceita um caminho que contenha uma planilha no formato .csv ou .xlsx')

    # Função abre a pasta de trabalho e a planilha (usada no construtor)
    def _abrir_pasta_trabalho(self):
        pasta_trabalho = openpyxl.load_workbook(self._caminho_planilha)
        nome_planilha = pasta_trabalho.get_sheet_names()[0]
        planilha = pasta_trabalho[nome_planilha]
        return pasta_trabalho, planilha

    # Função verifica se a planilha está no formato .csv e converte para .xlsx (usada no construtor)
    def _converter_csv_para_xlsx(self):
        if self._caminho_planilha.endswith('.csv'):
            # Abre a planilha no formato csv usando pandas
            planilha_csv = pd.read_csv(self._caminho_planilha, sep=';')
            # Nome da planilha que será salva no formato 'Nome da empresa' + .xlsx
            nome_planilha_excel = f'{self._nome_empresa.title()}.xlsx'
            # Caminho base da planilha csv
            caminho_base = os.path.split(self._caminho_planilha)[0]
            # Guarda o caminho que será salvo a planilha em formato xlsx
            caminho_planilha_excel = f'{caminho_base}\\{nome_planilha_excel}'
            # Salva em formato xlsx
            planilha_csv.to_excel(caminho_planilha_excel, index=False)
            # Deleta a planilha em formato csv
            os.remove(self._caminho_planilha)
            # Altera o caminho da planilha com o novo formato
            self._caminho_planilha = caminho_planilha_excel
        else:
            print('Planilha não está no formato .csv')

    # Função valida o parâmetro de coluna passado para as funções
    def _validar_coluna(self, parametro_coluna):
        # Primeira verificação é se for um inteiro
        if isinstance(parametro_coluna, int):
            # Verifica se o valor traz uma coluna que existe na planilha
            if parametro_coluna not in range(1, self._planilha.max_column + 1):
                raise ValueError('A coluna informada não está no intervalo da planilha!')
            else:
                return parametro_coluna
        elif isinstance(parametro_coluna, str):
            # Primeiro trata o string como se fosse o valor do cabeçalho
            # Verifica se o valor é encontrado no cabeçalho da planilha
            if self._encontrar_coluna_pelo_cabecalho(parametro_coluna) != 0:
                indice_coluna, letra_coluna = self._encontrar_coluna_pelo_cabecalho(parametro_coluna)
                return indice_coluna
            # Se não estiver no cabeçalho, tenta tratar o parâmetro como a letra que representa a coluna
            # Se a string não for composta apenas por letras ou seu tamanho for maior que 3, retorna um ValueError
            elif not parametro_coluna.isalpha() or len(parametro_coluna) > 3:
                raise ValueError('Valor de coluna inválido!')
            else:
                # Retorna o índice correspondente através das letras da coluna
                indice_coluna = column_index_from_string(parametro_coluna)
                # Verifica se o índice está no intervalo da planilha, se não estiver retorna um ValueError
                if indice_coluna not in range(1, self._planilha.max_column + 1):
                    raise ValueError('A coluna informada não está no intervalo da planilha!')
                else:
                    return indice_coluna

    # Função para definir o nome da planilha como data do dia + nome da empresa
    # O objetivo é salvar a planilha como histórico em um formato personalizado
    def _nome_personalizado_planilha(self):
        # Conseguindo a data de hoje
        data_hoje = date.today()
        # Extraindo a data de hoje no formato dd.MM.aaaa
        data_hoje_formatada = f'{data_hoje.day}.{data_hoje.month}.{data_hoje.year}'
        # Nome da empresa formatado
        nome_empresa_formatado = self._nome_empresa.title()
        # O nome a ser salvo será data formata + '-' + nome_empresa
        nome_personalizado = f'{data_hoje_formatada}-{nome_empresa_formatado}.xlsx'
        # Retorna o nome_personalizado
        return nome_personalizado

    # Função deixa a planilha em ordem alfabética de acordo com uma coluna
    def ordem_alfabetica_coluna(self, nome_coluna):
        # Lê a planilha
        planilha = pd.read_excel(self._caminho_planilha)
        # Ordena de acordo com a coluna passada como parâmetro
        planilha.sort_values(by=[nome_coluna], inplace=True)
        # Salva a planilha com as alterações realizadas
        planilha.to_excel(self._caminho_planilha, index=False)

    # Função dimensiona as colunas de acordo com os valores que estão contidos nelas
    def dimensionar_colunas(self, salvar=False):
        # Inicia a variável que representa o maior tamanho de texto da coluna
        maior_tamanho = 0
        # Percorre cada coluna da planilha
        for coluna in range(1, self._planilha.max_column + 1):
            # Salva a letra da coluna em uma variável
            letra_coluna = get_column_letter(coluna)
            # Percorre cada linha da coluna
            for linha in range(1, self._planilha.max_row + 1):
                # Verifica o tamanho do texto de cada célula
                tamanho_celula = len(str(self._planilha.cell(row=linha, column=coluna).value))
                # maior_tamanho vai assumir no final do laço o tamanho do maior texto
                if tamanho_celula > maior_tamanho:
                    maior_tamanho = tamanho_celula
            # Define a largura da coluna como o maior_tamanho + 3
            self._planilha.column_dimensions[letra_coluna].width = maior_tamanho + 3
            # Volta o valor de maior_tamanho para 0 para iniciar o próximo laço
            maior_tamanho = 0
            # Se o parâmetro salvar for passado como True, salva a planilha alterada
            if salvar:
                self.salvar_pasta_trabalho()

    # Função deleta uma sequência de colunas
    def deletar_colunas(self, primeira_coluna, quantidade_colunas_sequencia, salvar=False):
        # Validando os parâmetros passados:
        primeira_coluna = self._validar_coluna(primeira_coluna)
        if not isinstance(quantidade_colunas_sequencia, int):
            raise ValueError('Parâmetro quantidade_colunas_sequencia precisa ser do tipo int!')
        # Removendo as colunas de acordo com os parâmetros passados
        self._planilha.delete_cols(primeira_coluna, quantidade_colunas_sequencia)
        # Redimensionando as colunas que ficaram na planilha
        self.dimensionar_colunas()
        # Se o parâmetro salvar for passado como True, salva a planilha alterada
        if salvar:
            self.salvar_pasta_trabalho()

    # Função realiza uma soma dos valores da coluna passada como parâmetro
    def somar_valores(self, coluna_soma, salvar=False):
        # Valida a coluna passada como parâmetro
        coluna_soma = self._validar_coluna(coluna_soma)
        # Inicializa a variável que vai receber a soma dos valores
        soma = 0
        # Percorre as linhas da planilha
        for linha in range(2, self._planilha.max_row + 1):
            # Armazena o valor da célula em uma variável
            valor = float(self._planilha.cell(row=linha, column=coluna_soma).value)
            # Soma o valor da célula com a variável soma
            soma += valor
        # Deixa o valor da soma com duas casas decimais
        soma = round(soma, 2)
        # Insere o valor da soma na próxima linha em branco da coluna
        self._planilha.cell(row=self._planilha.max_row + 1, column=coluna_soma).value = soma
        # Se o parâmetro salvar for passado como True, salva a planilha alterada
        if salvar:
            self.salvar_pasta_trabalho()

    # Função retorna letra e índice de uma coluna de acordo com o cabeçalho da mesma
    def _encontrar_coluna_pelo_cabecalho(self, nome_cabecalho):
        # Percorre as colunas da planilha
        for coluna in range(1, self._planilha.max_column + 1):
            # Verifica se o cabeçalho desta coluna é o mesmo passado no parâmetro
            if self._planilha.cell(row=1, column=coluna).value == nome_cabecalho:
                # Armazena em uma variável a letra referente à coluna
                letra_coluna = get_column_letter(coluna)
                # Retorna o número e a letra da coluna que batem com o cabeçalho
                return coluna, letra_coluna
        # Retorna 0 no caso de não encontrar o valor do parâmetro
        return 0

    # Função salva a pasta de trabalho
    def salvar_pasta_trabalho(self, caminho=None, usar_nome_personalizado=False):
        if caminho is None:
            caminho = self._caminho_planilha

        if usar_nome_personalizado:
            # Extrai o caminho base onde será salva a planilha
            caminho_base = os.path.split(self._caminho_planilha)[0]
            # Retorna o nome personalizado no formato data-nome_empresa.xlsx
            nome_personalizado = self._nome_personalizado_planilha()
            # Concatena o caminho base com o nome personalizado
            caminho_personalizado = f'{caminho_base}\\{nome_personalizado}'
            # Salva com o nome personalizado
            self._pasta_trabalho.save(caminho_personalizado)
        else:
            # Salva com o mesmo nome
            self._pasta_trabalho.save(caminho)
